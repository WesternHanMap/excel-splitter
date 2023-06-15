# This is a function takes the pathname to the input table
import os

def splitTime(input_file):
    import pandas as pd
    import openpyxl
    import re
    import math

    # Get the directory path and filename of the input file
    input_dir, input_filename = os.path.split(input_file)
    # Create the output file path by prefixing "splited_" to the input filename
    output_filename = "splited_" + input_filename
    output_file = os.path.join(input_dir, output_filename)

    # Read the Excel file
    df = pd.read_excel(input_file)
    df = df[df.filter(regex='^Whether it is').notnull().any(axis=1)]
    # start time
    start_col = df.filter(regex=r'^Start').columns[0]
    new_column_name1 = 'start_time'
    df.rename(columns={start_col: new_column_name1}, inplace=True)

    # end time
    end_col = df.filter(regex=r'^End').columns[0]
    new_column_name2 = 'end_time'
    df.rename(columns={end_col: new_column_name2}, inplace=True)
    # Delete the last column
    df.drop(df.columns[-1], axis=1, inplace=True)

    # Function to parse the time string and convert it to a numeric value
    def parse_time(time_str):
        numeric_part = re.search(r'\d+', time_str)  # Extract the numeric part of the string
        if numeric_part is not None:
            numeric_value = int(numeric_part.group())
            if time_str.endswith('BCE'):
                return -numeric_value
            elif time_str.endswith('CE'):
                return numeric_value
        return None

    # Replace values in the 'Start time' column
    df['start_time'] = df['start_time'].apply(parse_time)

    # Replace values in the 'End time' column
    df['end_time'] = df['end_time'].apply(parse_time)
    key = sorted(set(df['start_time']).union(set(df['end_time'])))

    key_timestamps = sorted([x for x in key if not math.isnan(x)])

    # Define a function to convert the time values
    def convert_time(t):
        if t < 0:
            return f"{abs(t)} BCE"
        else:
            return f"{t} CE"

    df_copy = pd.read_excel(input_file)
    new_excel = pd.ExcelWriter(output_file, engine='xlsxwriter')
    df_copy.to_excel(new_excel, sheet_name='Original', index=False)

    # Iterate over the key timestamps and create sub-excel sheets
    for i in range(len(key_timestamps) - 1):
        start_time = key_timestamps[i]
        end_time = key_timestamps[i + 1]

        # Filter rows within the time period
        sub_df = df[(df['start_time'] <= start_time) & (df['end_time'] >= end_time)]

        # Write sub-excel sheet for the time period
        s = convert_time(int(start_time))
        t = convert_time(int(end_time))
        sheet_name = f'{s}-{t}'
        sub_df.to_excel(new_excel, sheet_name=sheet_name, index=False)

    new_excel.close()

    # Open the saved Excel file using openpyxl
    wb = openpyxl.load_workbook(output_file)

    # Iterate through each sheet
    for sheet_name in wb.sheetnames:
        # Select the current sheet
        sheet = wb[sheet_name]

        # Find the column indexes for 'start_time' and 'end_time'
        start_time_col = None
        end_time_col = None
        for col in sheet.iter_cols(min_col=1, max_col=sheet.max_column, min_row=1, max_row=1):
            for cell in col:
                if cell.value == 'start_time':
                    start_time_col = cell.column
                elif cell.value == 'end_time':
                    end_time_col = cell.column

        # Apply convert_time to 'start_time' and 'end_time' columns
        if start_time_col and end_time_col:
            for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=start_time_col, max_col=end_time_col):
                for cell in row:
                    cell.value = convert_time(cell.value)

    # Save the modified Excel file
    wb.save(output_file)

# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    directory = 'xlsx'
    # Iterate through the files in the directory
    for filename in os.listdir(directory):
        # Check if the filename does not contain "splited"
        if "splited" not in filename:
            splitTime(directory + "/" + filename)